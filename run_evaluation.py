import os, sys, csv, json
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt

try:
    import openpyxl
except ImportError:
    os.system(f"{sys.executable} -m pip install openpyxl --break-system-packages -q")
    import openpyxl

BASE_DIR    = os.path.join(os.path.expanduser("~"), "deadlift_trainer")
GT_PATH     = os.path.join(BASE_DIR, "ground_truth.xlsx")
RECORDINGS  = os.path.join(BASE_DIR, "data", "recordings")
EVAL_DIR    = os.path.join(BASE_DIR, "data", "evaluation")
os.makedirs(EVAL_DIR, exist_ok=True)

DARK  = '#1e1e1e'
PANEL = '#2b2b2b'
GREEN = '#50e880'
RED   = '#ff5555'
AMBER = '#ffaa00'
BLUE  = '#4488dd'
TEXT  = '#dddddd'
MUTED = '#888888'

plt.rcParams.update({
    'text.color':       TEXT,
    'axes.labelcolor':  MUTED,
    'xtick.color':      MUTED,
    'ytick.color':      MUTED,
    'axes.edgecolor':   '#555555',
    'figure.facecolor': DARK,
    'axes.facecolor':   DARK,
})


# ══════════════════════════════════════════════════════════════
#  LOAD EXCEL
# ══════════════════════════════════════════════════════════════

def load_excel():
    wb = openpyxl.load_workbook(GT_PATH)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]

    required = ["real_rounding","real_tilt","sys_rounding","sys_tilt","sys_score"]
    for col in required:
        if col not in headers:
            print(f"ERROR: Column '{col}' missing. Found: {headers}")
            sys.exit(1)

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        if any(d.get(k) is None for k in required):
            continue
        try:
            raw = d["sys_score"]
            if isinstance(raw, str):
                score = float(raw.replace('%','').strip())
                if score <= 1.0: score *= 100
            else:
                score = float(raw)
                if score <= 1.0: score *= 100

            rows.append({
                "session":      str(d.get("session_folder","")),
                "gt_rounding":  int(d["real_rounding"]),
                "gt_tilt":      int(d["real_tilt"]),
                "sys_rounding": int(d["sys_rounding"]),
                "sys_tilt":     int(d["sys_tilt"]),
                "sys_score":    score,
            })
        except: continue

    n_good = sum(1 for r in rows if r["gt_rounding"]==0 and r["gt_tilt"]==0)
    n_r    = sum(1 for r in rows if r["gt_rounding"]==1)
    n_t    = sum(1 for r in rows if r["gt_tilt"]==1)
    print(f"Loaded {len(rows)} labelled sessions from Excel")
    print(f"  Good form: {n_good}  |  Back rounding: {n_r}  |  Tilt: {n_t}")
    return rows


# ══════════════════════════════════════════════════════════════
#  METRICS
# ══════════════════════════════════════════════════════════════

def metrics(rows, pred_col, gt_col):
    tp = sum(1 for r in rows if r[pred_col]==1 and r[gt_col]==1)
    fp = sum(1 for r in rows if r[pred_col]==1 and r[gt_col]==0)
    tn = sum(1 for r in rows if r[pred_col]==0 and r[gt_col]==0)
    fn = sum(1 for r in rows if r[pred_col]==0 and r[gt_col]==1)
    n  = tp+fp+tn+fn
    pr = tp/(tp+fp) if (tp+fp)>0 else 0
    re = tp/(tp+fn) if (tp+fn)>0 else 0
    f1 = 2*pr*re/(pr+re) if (pr+re)>0 else 0
    ac = (tp+tn)/n if n>0 else 0
    return dict(tp=tp,fp=fp,tn=tn,fn=fn,n=n,
                precision=pr,recall=re,f1=f1,accuracy=ac)


# ══════════════════════════════════════════════════════════════
#  1. CONFUSION MATRIX — numbers only, no prose
# ══════════════════════════════════════════════════════════════

def plot_confusion_matrix(m, title, out_path):
    fig, ax = plt.subplots(figsize=(6, 5.5))
    fig.patch.set_facecolor(DARK)
    ax.set_facecolor(DARK)

    vals    = [[m["tn"], m["fp"]], [m["fn"], m["tp"]]]
    abbrevs = [["TN", "FP"], ["FN", "TP"]]
    cols    = [["#1d5c1d", "#7a1a1a"], ["#7a1a1a", "#1d5c1d"]]

    for i in range(2):
        for j in range(2):
            ax.add_patch(plt.Rectangle([j, 1-i], 1, 1,
                color=cols[i][j], alpha=0.92, zorder=1))
            # Big number
            ax.text(j+0.5, 1.5-i+0.1, str(vals[i][j]),
                ha='center', va='center',
                fontsize=44, fontweight='bold', color='white', zorder=2)
            # Small label
            ax.text(j+0.5, 1.5-i-0.22, abbrevs[i][j],
                ha='center', va='center',
                fontsize=13, color='#cccccc', zorder=2)

    ax.set_xlim(0, 2); ax.set_ylim(0, 2)

    # Column / row labels
    for j, lbl in enumerate(["Predicted: NO", "Predicted: YES"]):
        ax.text(j+0.5, 2.06, lbl, ha='center',
                fontsize=11, color=TEXT, fontweight='bold')
    for i, lbl in enumerate(["Actual: NO", "Actual: YES"]):
        ax.text(-0.03, 1.5-i, lbl, ha='right', va='center',
                fontsize=11, color=TEXT, fontweight='bold')

    ax.set_xticks([]); ax.set_yticks([])
    for sp in ax.spines.values(): sp.set_visible(False)

    ax.set_title(title, color='white', fontsize=13,
                 fontweight='bold', pad=32)

    # Single line of metrics at bottom
    line = (f"Precision {m['precision']:.0%}   "
            f"Recall {m['recall']:.0%}   "
            f"F1 {m['f1']:.0%}   "
            f"Accuracy {m['accuracy']:.0%}   "
            f"n = {m['n']}")
    fig.text(0.5, 0.03, line, ha='center', color=MUTED, fontsize=9)

    plt.tight_layout(rect=[0.06, 0.07, 1, 1])
    plt.savefig(out_path, dpi=150, bbox_inches='tight', facecolor=DARK)
    plt.close()
    print(f"  Saved: {os.path.basename(out_path)}")


# ══════════════════════════════════════════════════════════════
#  2. SCORE DISTRIBUTION — clean histogram, no arrows
# ══════════════════════════════════════════════════════════════

def plot_score_distribution(rows):
    good     = [r["sys_score"] for r in rows
                if r["gt_rounding"]==0 and r["gt_tilt"]==0]
    rounding = [r["sys_score"] for r in rows if r["gt_rounding"]==1]
    tilt     = [r["sys_score"] for r in rows
                if r["gt_tilt"]==1 and r["gt_rounding"]==0]

    fig, ax = plt.subplots(figsize=(10, 4.5))
    fig.patch.set_facecolor(DARK)
    ax.set_facecolor(DARK)
    bins = range(0, 106, 5)

    if good:
        ax.hist(good,     bins=bins, alpha=0.85, color=GREEN,
                label=f'Good form  (n={len(good)})',
                edgecolor=DARK, linewidth=0.5)
    if rounding:
        ax.hist(rounding, bins=bins, alpha=0.85, color=RED,
                label=f'Back rounding  (n={len(rounding)})',
                edgecolor=DARK, linewidth=0.5)
    if tilt:
        ax.hist(tilt,     bins=bins, alpha=0.85, color=AMBER,
                label=f'Lateral tilt  (n={len(tilt)})',
                edgecolor=DARK, linewidth=0.5)

    ax.axvline(85, color='white', ls='--', lw=1.5, alpha=0.5,
               label='Threshold (85%)')

    ax.set_xlabel('System form score (%)', fontsize=12)
    ax.set_ylabel('Count', fontsize=12)
    ax.set_title('Score distribution by ground truth class', fontsize=13,
                 color='white', pad=10)
    ax.set_xlim(0, 105)
    ax.legend(facecolor=PANEL, edgecolor='#555555',
              labelcolor=TEXT, fontsize=10)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)

    out = os.path.join(EVAL_DIR, "score_distribution.png")
    plt.tight_layout()
    plt.savefig(out, dpi=150, bbox_inches='tight', facecolor=DARK)
    plt.close()
    print(f"  Saved: {os.path.basename(out)}")


# ══════════════════════════════════════════════════════════════
#  3. FATIGUE ANALYSIS — reads session_summary.json directly
# ══════════════════════════════════════════════════════════════

def plot_fatigue_analysis():
    """
    Reads rep_scores arrays from every session_summary.json.
    Groups scores by rep position (1st, 2nd, 3rd...) across sessions.
    Shows whether form score changes across a set.
    """
    by_position = {}   # position (1,2,3...) -> list of scores

    if not os.path.isdir(RECORDINGS):
        print("  WARNING: recordings folder not found — skipping fatigue")
        return

    loaded = 0
    for d in os.listdir(RECORDINGS):
        sp = os.path.join(RECORDINGS, d, "session_summary.json")
        if not os.path.exists(sp):
            continue
        try:
            with open(sp) as f:
                data = json.load(f)
            rep_scores = data.get("rep_scores", [])
            if not rep_scores:
                continue
            for pos_idx, score in enumerate(rep_scores):
                pos = pos_idx + 1   # 1-indexed position
                if pos not in by_position:
                    by_position[pos] = []
                by_position[pos].append(float(score))
            loaded += 1
        except: continue

    print(f"  Fatigue: loaded rep scores from {loaded} sessions")

    if not by_position:
        print("  No rep score data found in session_summary.json files")
        return

    max_pos   = 12
    positions = sorted(k for k in by_position if k <= max_pos)
    # Only keep positions with at least 2 data points for meaningful stats
    positions = [p for p in positions if len(by_position[p]) >= 2]

    if len(positions) < 2:
        print("  Not enough data points for fatigue analysis "
              "(need ≥2 sessions with the same rep count)")
        # Still plot what we have
        positions = sorted(k for k in by_position if k <= max_pos)

    means  = [np.mean(by_position[p]) for p in positions]
    stds   = [np.std(by_position[p])  for p in positions]
    counts = [len(by_position[p])     for p in positions]

    fig, ax = plt.subplots(figsize=(11, 5))
    fig.patch.set_facecolor(DARK)
    ax.set_facecolor(DARK)

    ax.plot(positions, means, color=GREEN, lw=2.5, marker='o',
            markersize=8, markerfacecolor='white',
            markeredgecolor=GREEN, markeredgewidth=2,
            label='Mean score', zorder=3)
    ax.fill_between(positions,
        [max(0,  m-s) for m, s in zip(means, stds)],
        [min(105,m+s) for m, s in zip(means, stds)],
        color=GREEN, alpha=0.12, label='±1 SD')

    # n= labels above each point
    for x, m, n in zip(positions, means, counts):
        ax.text(x, min(m+stds[positions.index(x)]+3, 103),
                f'n={n}', ha='center', color=MUTED, fontsize=9)

    # Trend line
    if len(positions) >= 3:
        z    = np.polyfit(positions, means, 1)
        p_fn = np.poly1d(z)
        x_tr = np.linspace(min(positions), max(positions), 100)
        ax.plot(x_tr, p_fn(x_tr), color=AMBER, lw=1.5,
                ls='--', alpha=0.7,
                label=f'Trend  {z[0]:+.1f}%/rep')

    ax.axhline(80, color='white', ls=':', lw=1, alpha=0.25)
    ax.set_xlim(min(positions)-0.5, max(positions)+0.5)
    ax.set_ylim(0, 110)
    ax.set_xticks(positions)
    ax.set_xticklabels([f'Rep {p}' for p in positions], fontsize=10)
    ax.set_xlabel('Rep position within set', fontsize=12)
    ax.set_ylabel('Form score (%)', fontsize=12)
    ax.set_title('Form score by rep position — fatigue analysis',
                 color='white', fontsize=13, pad=10)
    ax.legend(facecolor=PANEL, edgecolor='#555555',
              labelcolor=TEXT, fontsize=10)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)

    out = os.path.join(EVAL_DIR, "fatigue_analysis.png")
    plt.tight_layout()
    plt.savefig(out, dpi=150, bbox_inches='tight', facecolor=DARK)
    plt.close()
    print(f"  Saved: {os.path.basename(out)}")

    if len(positions) >= 2:
        drop = means[0] - means[-1]
        print(f"  Rep 1 avg: {means[0]:.1f}%  →  "
              f"Rep {positions[-1]} avg: {means[-1]:.1f}%  "
              f"(change: {drop:+.1f}%)")


# ══════════════════════════════════════════════════════════════
#  4. 2D vs 3D — clean bars, no text boxes
# ══════════════════════════════════════════════════════════════

def plot_2d_vs_3d(rows):
    rows_2d = [{**r, "sys_tilt": 0} for r in rows]

    m_2d_r = metrics(rows_2d, "sys_rounding", "gt_rounding")
    m_2d_t = metrics(rows_2d, "sys_tilt",     "gt_tilt")
    m_3d_r = metrics(rows,    "sys_rounding", "gt_rounding")
    m_3d_t = metrics(rows,    "sys_tilt",     "gt_tilt")

    mkeys   = ['precision', 'recall', 'f1', 'accuracy']
    mlabels = ['Precision', 'Recall', 'F1-Score', 'Accuracy']
    x       = np.arange(len(mkeys))
    w       = 0.35

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(13, 5))
    fig.patch.set_facecolor(DARK)
    fig.suptitle('2D single-camera vs 3D dual-camera — detection performance',
                 color='white', fontsize=13, fontweight='bold')

    for ax, m2, m3, title in [
        (ax1, m_2d_r, m_3d_r, 'Back Rounding'),
        (ax2, m_2d_t, m_3d_t, 'Lateral Tilt'),
    ]:
        ax.set_facecolor(DARK)
        v2 = [m2[k] for k in mkeys]
        v3 = [m3[k] for k in mkeys]

        b2 = ax.bar(x - w/2, v2, w, label='2D (side only)',
                    color=BLUE,  alpha=0.88, edgecolor=DARK)
        b3 = ax.bar(x + w/2, v3, w, label='3D (dual-camera)',
                    color=GREEN, alpha=0.88, edgecolor=DARK)

        # Value labels above bars — no overlap risk
        for bar in list(b2) + list(b3):
            h = bar.get_height()
            if h >= 0.01:
                ax.text(bar.get_x() + bar.get_width()/2,
                        h + 0.025, f'{h:.0%}',
                        ha='center', va='bottom',
                        color=TEXT, fontsize=11, fontweight='bold')

        ax.set_ylim(0, 1.18)
        ax.set_xticks(x)
        ax.set_xticklabels(mlabels, fontsize=11)
        ax.set_ylabel('Score', fontsize=11)
        ax.set_title(title, color='white', fontsize=12, pad=8)
        ax.legend(facecolor=PANEL, edgecolor='#555555',
                  labelcolor=TEXT, fontsize=10)
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)

    plt.tight_layout()
    out = os.path.join(EVAL_DIR, "2d_vs_3d_comparison.png")
    plt.savefig(out, dpi=150, bbox_inches='tight', facecolor=DARK)
    plt.close()
    print(f"  Saved: {os.path.basename(out)}")
    return m_2d_r, m_2d_t, m_3d_r, m_3d_t


# ══════════════════════════════════════════════════════════════
#  RESULTS TABLE + CSV
# ══════════════════════════════════════════════════════════════

def save_results(m_r, m_t, m_2d_t, m_3d_t):
    print()
    print("═" * 62)
    print("  RESULTS SUMMARY")
    print("═" * 62)
    print(f"  {'Metric':<22} {'Back Rounding':>18} {'Lateral Tilt':>18}")
    print("  " + "─" * 60)
    for lbl, vr, vt in [
        ("n (reps evaluated)", m_r["n"],              m_t["n"]),
        ("True Positives",     m_r["tp"],             m_t["tp"]),
        ("False Positives",    m_r["fp"],             m_t["fp"]),
        ("True Negatives",     m_r["tn"],             m_t["tn"]),
        ("False Negatives",    m_r["fn"],             m_t["fn"]),
        ("──────────────────","──────────────────","──────────────────"),
        ("Precision",    f"{m_r['precision']:.1%}", f"{m_t['precision']:.1%}"),
        ("Recall",       f"{m_r['recall']:.1%}",    f"{m_t['recall']:.1%}"),
        ("F1-Score",     f"{m_r['f1']:.1%}",        f"{m_t['f1']:.1%}"),
        ("Accuracy",     f"{m_r['accuracy']:.1%}",  f"{m_t['accuracy']:.1%}"),
    ]:
        print(f"  {lbl:<22} {str(vr):>18} {str(vt):>18}")

    print()
    print("  2D vs 3D — Lateral Tilt")
    print(f"  {'':22} {'2D':>10} {'3D':>10}")
    print("  " + "─" * 44)
    for lbl, v2, v3 in [
        ("Precision", m_2d_t['precision'], m_3d_t['precision']),
        ("Recall",    m_2d_t['recall'],    m_3d_t['recall']),
        ("F1-Score",  m_2d_t['f1'],        m_3d_t['f1']),
    ]:
        print(f"  {lbl:<22} {v2:>9.1%} {v3:>9.1%}")
    print("═" * 62)

    out = os.path.join(EVAL_DIR, "results_summary.csv")
    with open(out, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["Metric", "Back Rounding", "Lateral Tilt"])
        for lbl, vr, vt in [
            ("n",        m_r["n"],                m_t["n"]),
            ("TP",       m_r["tp"],               m_t["tp"]),
            ("FP",       m_r["fp"],               m_t["fp"]),
            ("TN",       m_r["tn"],               m_t["tn"]),
            ("FN",       m_r["fn"],               m_t["fn"]),
            ("Precision",f"{m_r['precision']:.4f}",f"{m_t['precision']:.4f}"),
            ("Recall",   f"{m_r['recall']:.4f}",   f"{m_t['recall']:.4f}"),
            ("F1",       f"{m_r['f1']:.4f}",       f"{m_t['f1']:.4f}"),
            ("Accuracy", f"{m_r['accuracy']:.4f}", f"{m_t['accuracy']:.4f}"),
        ]:
            w.writerow([lbl, vr, vt])
    print(f"\n  CSV: {out}")


# ══════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════

if __name__ == "__main__":
    print("\n" + "═"*60)
    print("  DEADLIFT FORM ANALYSER — EVALUATION v3")
    print("═"*60 + "\n")

    rows = load_excel()
    if len(rows) < 3:
        print("Need at least 3 labelled rows."); sys.exit(1)

    print("\nGenerating charts...\n")

    m_r = metrics(rows, "sys_rounding", "gt_rounding")
    m_t = metrics(rows, "sys_tilt",     "gt_tilt")

    plot_confusion_matrix(m_r, "Back Rounding Detection",
        os.path.join(EVAL_DIR, "confusion_rounding.png"))
    plot_confusion_matrix(m_t, "Lateral Tilt Detection",
        os.path.join(EVAL_DIR, "confusion_tilt.png"))

    plot_score_distribution(rows)
    plot_fatigue_analysis()     # reads from session_summary.json files
    m_2d_r, m_2d_t, m_3d_r, m_3d_t = plot_2d_vs_3d(rows)
    save_results(m_r, m_t, m_2d_t, m_3d_t)

    print(f"\n  All outputs: {EVAL_DIR}\n")
